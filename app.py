"""
Compras Auto Peças — separa o relatório do ERP por destino de compra
(Embrepar / Rede / Fora) e compara cotações dos fornecedores.

Fase 1: Etapa A (separar) + Etapa B (comparar preços).
"""
import io
import re
import json
import zipfile
import sqlite3
import unicodedata
from datetime import datetime, date, timedelta
from pathlib import Path

import pandas as pd
import streamlit as st

# ----------------------------------------------------------------------------
# Configuração e dados fixos
# ----------------------------------------------------------------------------
BASE = Path(__file__).parent
MARCAS_PATH = BASE / "dados" / "marcas.json"
DB_PATH = BASE / "dados" / "historico.db"

st.set_page_config(page_title="Compras Auto Peças", page_icon="🔧", layout="wide")

# ----------------------------------------------------------------------------
# Visual (CSS)
# ----------------------------------------------------------------------------
def aplica_estilo():
    st.markdown(
        """
        <style>
        /* fundo geral */
        .stApp {
            background:
                radial-gradient(1200px 500px at 100% -10%, #1f2a44 0%, rgba(31,42,68,0) 55%),
                linear-gradient(180deg, #0f1320 0%, #11162400 100%),
                #0f1320;
            color: #e8ecf3;
        }
        .block-container { padding-top: 1.4rem; max-width: 1200px; }

        /* ---- legibilidade dos textos no fundo escuro ---- */
        .stApp, .stApp p, .stApp span, .stApp label, .stApp li,
        .stMarkdown, .stMarkdown p, .stCaption, [data-testid="stCaptionContainer"],
        .stRadio label, .stCheckbox label, h1, h2, h3, h4, h5, h6 {
            color: #eef2f8 !important;
        }
        /* rótulos de campos (uploader, inputs, number) */
        [data-testid="stWidgetLabel"] p,
        [data-testid="stWidgetLabel"] label,
        .stFileUploader label, .stTextInput label, .stNumberInput label {
            color: #f3f6fb !important; font-weight: 600;
        }
        /* texto auxiliar do uploader (limite, drag and drop) */
        [data-testid="stFileUploaderDropzone"] div,
        [data-testid="stFileUploaderDropzone"] span,
        [data-testid="stFileUploaderDropzoneInstructions"] * {
            color: #dde4ee !important;
        }
        /* legenda menor um pouco mais suave, mas ainda legível */
        .stCaption, [data-testid="stCaptionContainer"] p { color: #b9c2d4 !important; }
        /* inputs digitáveis */
        .stTextInput input, .stNumberInput input {
            color: #0f1320 !important; background:#eef2f8 !important;
        }
        /* tabelas/dataframe com texto claro */
        [data-testid="stDataFrame"] * { color: #0f1320; }
        /* caixas de aviso (info/warning/success) com texto escuro pra contrastar com o fundo claro delas */
        [data-testid="stNotification"] * { color: #1a1f2b !important; }
        /* expander: fundo escuro fixo + texto claro (não depende do estado do Streamlit) */
        [data-testid="stExpander"], [data-testid="stExpander"] details {
            background: rgba(255,255,255,.05) !important;
            border: 1px solid rgba(255,255,255,.12) !important;
            border-radius: 12px !important;
        }
        [data-testid="stExpander"] summary,
        [data-testid="stExpander"] summary p,
        [data-testid="stExpander"] summary span,
        [data-testid="stExpander"] [data-testid="stExpanderDetails"] *,
        details summary p, details summary span {
            color: #eef2f8 !important;
        }
        [data-testid="stExpander"] summary p { font-weight: 700 !important; }
        [data-testid="stExpander"] summary:hover p { color: #ffb53d !important; }
        /* fundo do cabeçalho do expander sempre escuro (Streamlit às vezes joga branco) */
        [data-testid="stExpander"] summary,
        [data-testid="stExpander"] details summary {
            background: #141a28 !important;
        }
        /* botões dentro do expander: texto escuro (não herdar o claro do conteúdo) */
        [data-testid="stExpanderDetails"] .stButton button,
        [data-testid="stExpanderDetails"] .stButton button *,
        [data-testid="stExpanderDetails"] .stDownloadButton button,
        [data-testid="stExpanderDetails"] .stDownloadButton button * {
            color: #1a1f2b !important;
        }
        /* botões claros (download/secundário) têm fundo branco -> texto escuro */
        .stButton button p, .stButton button span, .stButton button div,
        .stDownloadButton button p, .stDownloadButton button span, .stDownloadButton button div {
            color: #1a1f2b !important; font-weight: 700;
        }
        /* botões primários têm fundo com gradiente -> texto branco */
        .stButton button[kind="primary"] *, .stDownloadButton button[kind="primary"] * {
            color: #ffffff !important;
        }

        /* cabeçalho herói */
        .hero {
            background: linear-gradient(120deg, #ff6a00 0%, #ee0979 100%);
            border-radius: 20px;
            padding: 26px 32px;
            margin-bottom: 22px;
            box-shadow: 0 12px 35px rgba(255,106,0,.25);
        }
        .hero h1 { color:#fff; font-size: 2.0rem; margin:0; font-weight:800; letter-spacing:-.5px; }
        .hero p  { color:#fff; opacity:.92; margin:.35rem 0 0; font-size:1.02rem; }

        /* cartões de número */
        .card {
            background: rgba(255,255,255,.04);
            border: 1px solid rgba(255,255,255,.08);
            border-radius: 16px;
            padding: 18px 20px;
            text-align: center;
            backdrop-filter: blur(6px);
            transition: transform .15s ease, border-color .15s ease;
        }
        .card:hover { transform: translateY(-3px); border-color: rgba(255,255,255,.22); }
        .card .num { font-size: 2.4rem; font-weight: 800; line-height: 1; }
        .card .lbl { font-size: .82rem; text-transform: uppercase; letter-spacing: 1px; opacity:.75; margin-top:6px; }
        .c-emb .num { color:#4da3ff; }
        .c-rede .num { color:#ff5d6c; }
        .c-fora .num { color:#ffb53d; }
        .c-nc  .num { color:#b388ff; }

        /* abas */
        .stTabs [data-baseweb="tab-list"] { gap: 6px; }
        .stTabs [data-baseweb="tab"] {
            background: rgba(255,255,255,.04);
            border-radius: 10px 10px 0 0;
            padding: 10px 18px;
            font-weight: 600;
        }
        .stTabs [aria-selected="true"] {
            background: linear-gradient(120deg, #ff6a00, #ee0979);
            color:#fff !important;
        }

        /* botões */
        .stDownloadButton button, .stButton button {
            border-radius: 12px;
            font-weight: 700;
            border: 1px solid rgba(255,255,255,.12);
            padding: .6rem 1rem;
            transition: transform .12s ease, box-shadow .12s ease;
        }
        /* botões claros (secundários): fundo claro fixo + texto escuro, inclusive no hover */
        .stDownloadButton button:not([kind="primary"]),
        .stButton button:not([kind="primary"]) {
            background: #f3f6fb !important;
            color: #1a1f2b !important;
        }
        .stDownloadButton button:not([kind="primary"]):hover,
        .stButton button:not([kind="primary"]):hover,
        .stDownloadButton button:not([kind="primary"]):focus,
        .stButton button:not([kind="primary"]):focus,
        .stDownloadButton button:not([kind="primary"]):active,
        .stButton button:not([kind="primary"]):active {
            background: #ffffff !important;
            color: #1a1f2b !important;
            transform: translateY(-2px);
            box-shadow: 0 8px 20px rgba(0,0,0,.35);
            border-color: #ff6a00 !important;
        }
        /* botões primários têm fundo com gradiente -> texto branco, inclusive hover */
        .stButton button[kind="primary"], .stDownloadButton button[kind="primary"],
        .stButton button[kind="primary"]:hover, .stDownloadButton button[kind="primary"]:hover {
            background: linear-gradient(120deg, #ff6a00, #ee0979) !important;
            border: none; color:#fff !important;
            transform: translateY(-2px);
        }

        /* uploader */
        [data-testid="stFileUploaderDropzone"] {
            background: rgba(255,255,255,.03);
            border: 1.5px dashed rgba(255,255,255,.22);
            border-radius: 14px;
        }
        /* botão Upload/Browse do uploader: fundo claro fixo + texto escuro (não some) */
        [data-testid="stFileUploader"] button,
        [data-testid="stFileUploaderDropzone"] button,
        [data-testid="stBaseButton-secondary"] {
            background: #f3f6fb !important;
            color: #1a1f2b !important;
            border: 1px solid rgba(0,0,0,.18) !important;
            font-weight: 700 !important;
        }
        [data-testid="stFileUploader"] button *,
        [data-testid="stFileUploaderDropzone"] button * {
            color: #1a1f2b !important;
        }
        [data-testid="stFileUploader"] button:hover,
        [data-testid="stFileUploaderDropzone"] button:hover {
            background: #ffffff !important;
            color: #1a1f2b !important;
            border-color: #ff6a00 !important;
        }
        /* tabelas */
        [data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; }
        </style>
        """,
        unsafe_allow_html=True,
    )


def cartao(coluna, classe, numero, rotulo):
    coluna.markdown(
        f"<div class='card {classe}'><div class='num'>{numero}</div>"
        f"<div class='lbl'>{rotulo}</div></div>",
        unsafe_allow_html=True,
    )


@st.cache_data
def carregar_marcas():
    """Lê a lista fixa de marcas por categoria e devolve um mapa MARCA->CATEGORIA."""
    cat = json.loads(MARCAS_PATH.read_text(encoding="utf-8"))
    mapa = {}
    for categoria, marcas in cat.items():
        for m in marcas:
            mapa[normaliza(m)] = categoria
    return cat, mapa


def normaliza(texto):
    """Deixa o texto comparável: maiúsculo, sem acento e sem espaços sobrando."""
    if texto is None:
        return ""
    s = str(texto).strip().upper()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s


def so_numero_qtd(valor):
    """'2 UN' -> 2. Pega o primeiro número inteiro que aparecer."""
    if valor is None:
        return 0
    m = re.search(r"\d+", str(valor))
    return int(m.group()) if m else 0


def para_preco(valor):
    """Converte 'R$ 1.234,56' / '1234,56' / '12.5' -> float. Vazio -> None."""
    if valor is None:
        return None
    s = str(valor).strip()
    if s == "":
        return None
    s = s.replace("R$", "").replace(" ", "")
    # remove separador de milhar e troca vírgula decimal por ponto
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        v = float(s)
        return v if v > 0 else None
    except ValueError:
        return None


def acha_coluna(df, *candidatos):
    """Encontra a coluna do DataFrame pelo nome (ignora acento/maiúscula)."""
    norm = {normaliza(c): c for c in df.columns}
    for cand in candidatos:
        if normaliza(cand) in norm:
            return norm[normaliza(cand)]
    return None


# ----------------------------------------------------------------------------
# Leitura do relatório do ERP
# ----------------------------------------------------------------------------
def ler_erp(arquivo):
    df = pd.read_excel(arquivo)
    col_codigo = acha_coluna(df, "Código", "Codigo", "Cod")
    col_produto = acha_coluna(df, "Produto", "Descrição", "Descricao")
    col_qtd = acha_coluna(df, "Sugestão", "Sugestao", "Qtd", "Quantidade")
    col_marca = acha_coluna(df, "Marca", "Fabricante")
    col_estoque = acha_coluna(df, "Estoque")
    col_situacao = acha_coluna(df, "Situação", "Situacao")

    faltando = [nome for nome, col in [
        ("Código", col_codigo), ("Produto", col_produto),
        ("Sugestão", col_qtd), ("Marca", col_marca)] if col is None]
    if faltando:
        raise ValueError(
            "A planilha do ERP não tem as colunas: " + ", ".join(faltando) +
            ". Colunas encontradas: " + ", ".join(str(c) for c in df.columns))

    out = pd.DataFrame({
        "Código": df[col_codigo].astype(str).str.strip(),
        "Produto": df[col_produto].astype(str).str.strip(),
        "Marca": df[col_marca].astype(str).str.strip(),
        "Quantidade": df[col_qtd].apply(so_numero_qtd),
    })
    if col_estoque:
        out["Estoque"] = df[col_estoque].astype(str).str.strip()
    if col_situacao:
        out["Situação"] = df[col_situacao].astype(str).str.strip()
    # tira linhas totalmente vazias
    out = out[out["Código"].str.len() > 0].reset_index(drop=True)
    return out


def ler_bloco_embrepar(arquivo):
    """Lê uma das planilhas da Embrepar (POA/Pelotas/Falta) e devolve lista de itens.
    Cada arquivo tem layout um pouco diferente, então acha o cabeçalho sozinho."""
    import openpyxl
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, hdr = None, None
    for i, r in enumerate(rows):
        cels = [str(c).strip().lower() if c is not None else "" for c in r]
        if "produto" in cels or "código" in cels or "codigo" in cels:
            hdr_idx, hdr = i, cels
            break
    if hdr_idx is None:
        raise ValueError("Não achei o cabeçalho (com 'Produto' ou 'Código') nessa planilha.")

    def idx(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return None

    ic = idx("produto", "código", "codigo")
    im = idx("marca")
    iq = idx("quant.", "quant", "sugestão", "sugestao", "quantidade")
    ip = idx("preço unit.", "preço unit", "preco unit.", "preço", "preco")

    itens = []
    for r in rows[hdr_idx + 1:]:
        cod = r[ic] if ic is not None else None
        if cod is None or str(cod).strip() == "":
            continue
        itens.append({
            "Código": str(cod).strip(),
            "Marca": str(r[im]).strip() if im is not None and r[im] is not None else "",
            "Quantidade": so_numero_qtd(r[iq]) if iq is not None else 0,
            "Preço Embrepar (R$)": para_preco(r[ip]) if ip is not None else None,
        })
    return itens


def condensar_embrepar(arq_poa, arq_pelotas, arq_falta):
    """Junta POA + PELOTAS + EM FALTA num único arquivo, empilhados, com
    linha-título separando cada bloco e coluna em branco pro concorrente cobrir."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    blocos = []
    contagem = {}
    for rotulo, arq, cor in [
        ("POA", arq_poa, "1F6FEB"),
        ("PELOTAS", arq_pelotas, "8957E5"),
        ("EM FALTA", arq_falta, "D29922"),
    ]:
        itens = ler_bloco_embrepar(arq) if arq is not None else []
        contagem[rotulo] = len(itens)
        blocos.append((rotulo, cor, itens))

    cols = ["Código", "Marca", "Quantidade", "Preço Referência (R$)", "PREÇO CONCORRENTE (R$)"]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotacao Concorrentes"

    borda = Border(*(Side(style="thin", color="DDDDDD"),) * 4)
    # cabeçalho de colunas
    ws.append(cols)
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="333333")
        c.alignment = Alignment(horizontal="center")
        c.border = borda

    for rotulo, cor, itens in blocos:
        # linha-título do bloco (mesclada)
        r = ws.max_row + 1
        ws.cell(row=r, column=1, value=f"{rotulo}  ({len(itens)} itens)")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(cols))
        cel = ws.cell(row=r, column=1)
        cel.font = Font(bold=True, color="FFFFFF", size=12)
        cel.fill = PatternFill("solid", fgColor=cor)
        cel.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        # itens
        for it in itens:
            ws.append([
                it["Código"], it["Marca"], it["Quantidade"],
                it["Preço Embrepar (R$)"], "",
            ])
            for c in ws[ws.max_row]:
                c.border = borda

    # larguras
    larguras = [22, 16, 12, 18, 22]
    for i, w in enumerate(larguras, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), contagem


def separar_por_destino(df_erp, mapa):
    """Adiciona a coluna Destino com base na marca."""
    df = df_erp.copy()
    df["Destino"] = df["Marca"].apply(lambda m: mapa.get(normaliza(m), "NÃO CLASSIFICADO"))
    return df


# ----------------------------------------------------------------------------
# Geração de Excel (download)
# ----------------------------------------------------------------------------
def excel_bytes(abas: dict):
    """abas = {nome_aba: dataframe}. Devolve bytes de um .xlsx."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nome, df in abas.items():
            df.to_excel(writer, sheet_name=nome[:31], index=False)
    buf.seek(0)
    return buf.getvalue()


def excel_relatorio(abas, resumo_sheet, nomes):
    """Igual ao excel_bytes, mas na aba de resumo pinta o MENOR preço de verde
    e o MAIOR de vermelho em cada linha (entre as colunas dos fornecedores)."""
    from openpyxl.styles import PatternFill
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for nome, df in abas.items():
            df.to_excel(writer, sheet_name=nome[:31], index=False)
        ws = writer.sheets[resumo_sheet[:31]]
        df_res = abas[resumo_sheet]
        cols = list(df_res.columns)
        idxs = {n: cols.index(n) + 1 for n in nomes if n in cols}
        verde = PatternFill("solid", fgColor="C6EFCE")
        vermelho = PatternFill("solid", fgColor="FFC7CE")
        for ri in range(len(df_res)):
            valores = {}
            for n, ci in idxs.items():
                v = df_res.iloc[ri][n]
                if pd.notna(v):
                    valores[ci] = float(v)
            if not valores:
                continue
            vmin, vmax = min(valores.values()), max(valores.values())
            for ci, v in valores.items():
                cell = ws.cell(row=ri + 2, column=ci)
                if v == vmin:
                    cell.fill = verde
                elif v == vmax:
                    cell.fill = vermelho
    buf.seek(0)
    return buf.getvalue()


def zip_pedidos(resumo, nomes):
    """Gera um ZIP com um arquivo de pedido (.xlsx) por fornecedor vencedor,
    cada um com os itens que ele ganhou + preço acordado + total."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for nome in nomes:
            d = resumo[resumo["Vencedor"] == nome]
            if d.empty:
                continue
            pedido = d[["Código", "Produto", "Marca", "Quantidade", "Preço unit", "Total"]].rename(
                columns={"Preço unit": "Preço"})
            nome_limpo = re.sub(r"[^A-Za-z0-9_-]+", "_", str(nome)).strip("_") or "fornecedor"
            z.writestr(f"pedido_{nome_limpo}.xlsx", excel_bytes({"Pedido": pedido}))
    buf.seek(0)
    return buf.getvalue()


# ----------------------------------------------------------------------------
# Banco de dados (histórico de cotações)
# ----------------------------------------------------------------------------
def conectar():
    con = sqlite3.connect(DB_PATH)
    con.execute("""CREATE TABLE IF NOT EXISTS cotacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        data TEXT NOT NULL,
        criado_em TEXT NOT NULL
    )""")
    con.execute("""CREATE TABLE IF NOT EXISTS itens (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cotacao_id INTEGER NOT NULL,
        codigo TEXT, produto TEXT, marca TEXT,
        quantidade INTEGER, vencedor TEXT,
        preco_unit REAL, total REAL
    )""")
    con.execute("""CREATE TABLE IF NOT EXISTS precos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cotacao_id INTEGER NOT NULL,
        codigo TEXT, fornecedor TEXT, preco REAL
    )""")
    return con


def salvar_cotacao(data_cot, resumo, nomes):
    """Grava uma rodada de cotação no histórico."""
    con = conectar()
    cur = con.cursor()
    cur.execute("INSERT INTO cotacoes (data, criado_em) VALUES (?, ?)",
                (data_cot.isoformat(), datetime.now().isoformat(timespec="seconds")))
    cid = cur.lastrowid
    for _, r in resumo.iterrows():
        cur.execute(
            "INSERT INTO itens (cotacao_id, codigo, produto, marca, quantidade, vencedor, preco_unit, total)"
            " VALUES (?,?,?,?,?,?,?,?)",
            (cid, str(r["Código"]), str(r["Produto"]), str(r["Marca"]),
             int(r["Quantidade"]), str(r["Vencedor"]),
             float(r["Preço unit"]), float(r["Total"])))
        for nome in nomes:
            p = r.get(nome)
            if pd.notna(p) and p is not None:
                cur.execute("INSERT INTO precos (cotacao_id, codigo, fornecedor, preco) VALUES (?,?,?,?)",
                            (cid, str(r["Código"]), nome, float(p)))
    con.commit()
    con.close()
    return cid


def ler_cotacoes():
    con = conectar()
    df = pd.read_sql_query("SELECT * FROM cotacoes ORDER BY data", con)
    con.close()
    if not df.empty:
        df["data"] = pd.to_datetime(df["data"]).dt.date
    return df


def ler_itens(ids=None):
    con = conectar()
    q = "SELECT i.*, c.data FROM itens i JOIN cotacoes c ON c.id=i.cotacao_id"
    if ids:
        q += " WHERE i.cotacao_id IN (%s)" % ",".join("?" * len(ids))
        df = pd.read_sql_query(q, con, params=ids)
    else:
        df = pd.read_sql_query(q, con)
    con.close()
    if not df.empty:
        df["data"] = pd.to_datetime(df["data"]).dt.date
    return df


def apagar_cotacao(cid):
    con = conectar()
    con.execute("DELETE FROM itens WHERE cotacao_id=?", (cid,))
    con.execute("DELETE FROM precos WHERE cotacao_id=?", (cid,))
    con.execute("DELETE FROM cotacoes WHERE id=?", (cid,))
    con.commit()
    con.close()


def moeda(v):
    """Formata número como R$ 1.234,56."""
    try:
        return "R$ " + f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except (ValueError, TypeError):
        return "R$ 0,00"


# ----------------------------------------------------------------------------
# Interface
# ----------------------------------------------------------------------------
categorias, mapa = carregar_marcas()

aplica_estilo()
st.markdown(
    "<div class='hero'><h1>🔧 Compras Auto Peças</h1>"
    "<p>Separe o relatório do ERP por fornecedor e descubra onde comprar mais barato.</p></div>",
    unsafe_allow_html=True,
)


# ----------------------------------------------------------------------------
# Trava de senha
# ----------------------------------------------------------------------------
def senha_correta():
    try:
        return st.secrets["app_password"]
    except Exception:
        return "autopecas"  # senha padrão só pra teste local


def checar_senha():
    if st.session_state.get("autenticado"):
        return True
    st.markdown("### 🔒 Acesso restrito")
    st.write("Digite a senha pra usar o sistema (aperte Enter ou clique em Entrar).")
    with st.form("login_form"):
        senha = st.text_input("Senha", type="password")
        enviar = st.form_submit_button("Entrar", type="primary")
    if enviar:
        if senha == senha_correta():
            st.session_state["autenticado"] = True
            st.rerun()
        else:
            st.error("Senha incorreta. Tente de novo.")
    return False


if not checar_senha():
    st.stop()

aba_a, aba_cond, aba_b, aba_dash, aba_cfg = st.tabs([
    "1️⃣ Separar relatório (Etapa A)",
    "🧲 Condensador Embrepar",
    "2️⃣ Comparar cotações (Etapa B)",
    "📊 Dashboard",
    "⚙️ Marcas cadastradas",
])

# ---------------- ETAPA A ----------------
with aba_a:
    st.subheader("Etapa A — Jogue o relatório do ERP")
    st.write("O app separa cada item por onde ele deve ser comprado.")
    arq = st.file_uploader("Relatório do ERP (.xlsx)", type=["xlsx", "xls"], key="erp")

    if arq is not None:
        try:
            erp = ler_erp(arq)
            dados = separar_por_destino(erp, mapa)
        except Exception as e:
            st.error(f"Não consegui ler a planilha: {e}")
            st.stop()

        emb = dados[dados["Destino"] == "EMBREPAR"]
        rede = dados[dados["Destino"] == "REDE"]
        fora = dados[dados["Destino"] == "FORA"]
        nc = dados[dados["Destino"] == "NÃO CLASSIFICADO"]

        c1, c2, c3, c4 = st.columns(4)
        cartao(c1, "c-emb", len(emb), "Embrepar")
        cartao(c2, "c-rede", len(rede), "Rede")
        cartao(c3, "c-fora", len(fora), "Cotação FORA")
        cartao(c4, "c-nc", len(nc), "Não classificados")
        st.write("")

        if len(nc) > 0:
            st.warning("Tem itens com marca fora da lista — confira na aba 'Não classificados' do arquivo, ou cadastre a marca em ⚙️.")
            st.dataframe(nc[["Código", "Produto", "Marca"]], use_container_width=True, hide_index=True)

        st.divider()
        st.write("**Baixe os resultados:**")

        cols_compra = ["Código", "Produto", "Marca", "Quantidade"]
        # 1) Arquivo SÓ Embrepar  e  2) arquivo SÓ Rede (separados)
        arq_embrepar = excel_bytes({"Comprar EMBREPAR": emb[cols_compra]})
        arq_rede = excel_bytes({"Comprar REDE": rede[cols_compra]})
        # 3) Cotação em branco pra mandar pros fornecedores
        cot = fora[cols_compra].copy()
        cot["PREÇO UNIT (R$)"] = ""
        cotacao = excel_bytes({"Cotacao": cot})

        d1, d2 = st.columns(2)
        d1.download_button(
            "🟦 Compra EMBREPAR",
            data=arq_embrepar, file_name="compra_embrepar.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)
        d2.download_button(
            "🟥 Compra REDE",
            data=arq_rede, file_name="compra_rede.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

        st.download_button(
            "📄 Cotação FORA (negociar em grande escala)",
            data=cotacao, file_name="cotacao_fora.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True)

        if len(nc) > 0:
            arq_nc = excel_bytes({"Nao classificados": nc[cols_compra]})
            st.download_button(
                f"⚠️ Não classificados ({len(nc)})",
                data=arq_nc, file_name="nao_classificados.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        st.info("📄 A **Cotação FORA** é a lista dos itens que os compradores negociam em grande escala (compra de lote). "
                "Para a **Embrepar**, use a aba **🧲 Condensador**.")

# ---------------- CONDENSADOR EMBREPAR ----------------
with aba_cond:
    st.subheader("🧲 Condensador Embrepar")
    st.write(
        "A Embrepar te devolve **3 planilhas** (POA, Pelotas e Em falta). "
        "Suba as 3 aqui que o app junta tudo em **um arquivo só**, empilhado nesta ordem: "
        "**POA → Pelotas → Em falta** — pronto pra mandar pros concorrentes cobrirem o preço."
    )
    cc1, cc2, cc3 = st.columns(3)
    f_poa = cc1.file_uploader("🔵 POA", type=["xlsx", "xls"], key="cond_poa")
    f_pel = cc2.file_uploader("🟣 PELOTAS", type=["xlsx", "xls"], key="cond_pel")
    f_fal = cc3.file_uploader("🟡 EM FALTA", type=["xlsx", "xls"], key="cond_fal")

    if st.button("🧲 Condensar em uma planilha só", type="primary"):
        if not any([f_poa, f_pel, f_fal]):
            st.warning("Suba pelo menos uma planilha.")
        else:
            try:
                arquivo, cont = condensar_embrepar(f_poa, f_pel, f_fal)
            except Exception as e:
                st.error(f"Não consegui ler uma das planilhas: {e}")
                st.stop()

            k1, k2, k3, k4 = st.columns(4)
            cartao(k1, "c-emb", cont.get("POA", 0), "POA")
            cartao(k2, "c-nc", cont.get("PELOTAS", 0), "Pelotas")
            cartao(k3, "c-fora", cont.get("EM FALTA", 0), "Em falta")
            cartao(k4, "c-rede", sum(cont.values()), "Total")
            st.write("")
            st.success("Planilha condensada gerada! Baixe abaixo e mande pros concorrentes.")
            st.download_button(
                "📥 Baixar planilha condensada",
                data=arquivo, file_name="cotacao_concorrentes.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary")


# ---------------- ETAPA B ----------------
with aba_b:
    st.subheader("Etapa B — Compare os preços dos fornecedores")
    st.write("Suba as cotações que cada fornecedor te devolveu preenchidas. O app acha o mais barato de cada item.")

    n = st.number_input("Quantos fornecedores responderam?", min_value=1, max_value=10, value=3, step=1)

    cotas = []
    for i in range(int(n)):
        c1, c2 = st.columns([1, 2])
        nome = c1.text_input(f"Nome do fornecedor {i+1}", value=f"Fornecedor {i+1}", key=f"nome_{i}")
        arqf = c2.file_uploader(f"Cotação preenchida — {nome}", type=["xlsx", "xls"], key=f"cot_{i}")
        if arqf is not None:
            cotas.append((nome, arqf))

    if cotas and st.button("🔍 Comparar preços", type="primary"):
        # lê cada cotação -> {codigo: preço}  (resultado vai pro session_state)
        precos = {}          # codigo -> {fornecedor: preço}
        info = {}            # codigo -> (produto, marca, qtd)
        nomes = []
        for nome, arqf in cotas:
            nomes.append(nome)
            df = pd.read_excel(arqf)
            cod = acha_coluna(df, "Código", "Codigo")
            prod = acha_coluna(df, "Produto")
            marca = acha_coluna(df, "Marca")
            qtd = acha_coluna(df, "Quantidade", "Qtd")
            preco = acha_coluna(df, "PREÇO CONCORRENTE (R$)", "Preço concorrente",
                                "PREÇO UNIT (R$)", "Preço unit", "Preço", "Preco", "Valor")
            ref = acha_coluna(df, "Preço Referência (R$)", "Preço referencia",
                              "Preço Embrepar (R$)", "Preço Embrepar")
            if cod is None or preco is None:
                st.error(f"A cotação de '{nome}' precisa ter as colunas 'Código' e 'PREÇO CONCORRENTE (R$)' "
                         "(ou 'PREÇO UNIT (R$)').")
                st.stop()
            for _, r in df.iterrows():
                codigo = str(r[cod]).strip()
                if not codigo or codigo.lower() == "nan":
                    continue
                # pula as linhas-título do condensador (POA / PELOTAS / EM FALTA)
                if "itens)" in codigo.lower() or codigo.upper() in ("POA", "PELOTAS", "EM FALTA"):
                    continue
                p = para_preco(r[preco])
                precos.setdefault(codigo, {})
                if p is not None:
                    precos[codigo][nome] = p
                # preço de referência da Embrepar entra na disputa (pregão)
                if ref is not None:
                    pr = para_preco(r[ref])
                    if pr is not None:
                        precos[codigo]["Embrepar"] = pr
                if codigo not in info:
                    info[codigo] = (
                        str(r[prod]).strip() if prod else "",
                        str(r[marca]).strip() if marca else "",
                        so_numero_qtd(r[qtd]) if qtd else 0,
                    )

        # se algum item tem preço de referência, a Embrepar vira um "concorrente" na disputa
        if any("Embrepar" in pmap for pmap in precos.values()) and "Embrepar" not in nomes:
            nomes.append("Embrepar")

        # monta resumo
        linhas = []
        sem_cotacao = []
        for codigo, pmap in precos.items():
            produto, marca, qtd = info.get(codigo, ("", "", 0))
            base = {"Código": codigo, "Produto": produto, "Marca": marca, "Quantidade": qtd}
            for nome in nomes:
                base[nome] = pmap.get(nome, None)
            if pmap:
                vencedor = min(pmap, key=pmap.get)
                preco_v = pmap[vencedor]
                base["Vencedor"] = vencedor
                base["Preço unit"] = preco_v
                base["Total"] = round(preco_v * qtd, 2)
                linhas.append(base)
            else:
                sem_cotacao.append(base)

        if not linhas:
            st.warning("Nenhum preço foi encontrado nas planilhas. Confira se a coluna de preço foi preenchida.")
            st.stop()

        resumo = pd.DataFrame(linhas)
        # ordena por marca/produto
        resumo = resumo.sort_values(["Vencedor", "Produto"]).reset_index(drop=True)
        # guarda na sessão pra não perder ao clicar em salvar/baixar
        st.session_state["resultado_b"] = {
            "resumo": resumo, "nomes": nomes, "sem_cotacao": sem_cotacao,
        }

    # ----- mostra o resultado (se já existir) -----
    if "resultado_b" in st.session_state:
        r = st.session_state["resultado_b"]
        resumo, nomes, sem_cotacao = r["resumo"], r["nomes"], r["sem_cotacao"]

        st.success(f"Comparei {len(resumo)} itens.")

        # totais por fornecedor
        tot = resumo.groupby("Vencedor")["Total"].agg(["count", "sum"]).reset_index()
        tot.columns = ["Fornecedor", "Itens", "Total R$"]
        st.write("**Quanto você vai comprar de cada um:**")
        st.dataframe(tot, use_container_width=True, hide_index=True)

        st.write("**Resumo (menor preço em destaque):**")
        st.dataframe(resumo, use_container_width=True, hide_index=True)

        # monta arquivo final: 1 aba resumo + 1 aba por fornecedor + sem cotação
        abas = {"Resumo menor preço": resumo}
        for nome in nomes:
            comprar = resumo[resumo["Vencedor"] == nome][
                ["Código", "Produto", "Marca", "Quantidade", "Preço unit", "Total"]]
            abas[nome] = comprar
        if sem_cotacao:
            abas["Ninguem cotou"] = pd.DataFrame(sem_cotacao)[["Código", "Produto", "Marca", "Quantidade"]]
            st.warning(f"{len(sem_cotacao)} item(ns) ninguém cotou — veja a aba 'Ninguem cotou' no arquivo.")

        cbaixa, czip = st.columns(2)
        cbaixa.download_button(
            "📥 Baixar relatório completo",
            data=excel_relatorio(abas, "Resumo menor preço", nomes),
            file_name="resultado_cotacao.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary", use_container_width=True)
        czip.download_button(
            "📦 Baixar pedidos por fornecedor (ZIP)",
            data=zip_pedidos(resumo, nomes), file_name="pedidos_por_fornecedor.zip",
            mime="application/zip", type="primary", use_container_width=True)
        st.caption("📦 O ZIP traz um arquivo de pedido por fornecedor (e da Embrepar), "
                   "só com o que cada um ganhou — é só mandar direto pra ele.")

        with st.expander("💾 Salvar no histórico (Dashboard)"):
            data_cot = st.date_input("Data desta cotação", value=date.today(),
                                     format="DD/MM/YYYY", key="data_salvar")
            if st.button("💾 Salvar no histórico"):
                salvar_cotacao(data_cot, resumo, nomes)
                st.success(f"Cotação de {data_cot.strftime('%d/%m/%Y')} salva! Veja na aba 📊 Dashboard.")

# ---------------- DASHBOARD ----------------
with aba_dash:
    st.subheader("📊 Dashboard de compras")
    cot = ler_cotacoes()

    if cot.empty:
        st.info("Ainda não tem nada no histórico. Vá na **Etapa B**, compare uma cotação e clique em **💾 Salvar no histórico**.")
    else:
        dmin, dmax = cot["data"].min(), cot["data"].max()
        # ----- filtro de calendário -----
        f1, f2 = st.columns([2, 1])
        periodo = f1.date_input(
            "📅 Período", value=(dmin, dmax), min_value=dmin, max_value=dmax,
            format="DD/MM/YYYY", key="dash_periodo")
        if isinstance(periodo, (list, tuple)) and len(periodo) == 2:
            ini, fim = periodo
        else:
            ini, fim = dmin, dmax

        cot_f = cot[(cot["data"] >= ini) & (cot["data"] <= fim)]
        itens = ler_itens(list(cot_f["id"]))

        if itens.empty:
            st.warning("Nenhuma cotação salva nesse período.")
        else:
            forn_sel = f2.multiselect(
                "Fornecedores", sorted(itens["vencedor"].unique()),
                default=sorted(itens["vencedor"].unique()), key="dash_forn")
            itens = itens[itens["vencedor"].isin(forn_sel)]

            # ----- KPIs -----
            total_geral = itens["total"].sum()
            n_cot = cot_f["id"].nunique()
            n_itens = len(itens)
            ticket = total_geral / n_cot if n_cot else 0
            k1, k2, k3, k4 = st.columns(4)
            cartao(k1, "c-fora", moeda(total_geral), "Total comprado")
            cartao(k2, "c-emb", n_cot, "Cotações")
            cartao(k3, "c-rede", n_itens, "Itens comprados")
            cartao(k4, "c-nc", moeda(ticket), "Média por cotação")
            st.write("")

            # ----- gasto ao longo do tempo -----
            st.markdown("#### 💰 Gasto por data")
            por_data = itens.groupby("data")["total"].sum().reset_index()
            por_data = por_data.rename(columns={"total": "Total R$", "data": "Data"}).set_index("Data")
            st.line_chart(por_data, height=260, color="#ffb53d")

            # ----- gasto por fornecedor -----
            cg1, cg2 = st.columns(2)
            with cg1:
                st.markdown("#### 🏪 Total por fornecedor")
                pf = itens.groupby("vencedor")["total"].sum().sort_values(ascending=False)
                pf = pf.rename_axis("Fornecedor").rename("Total R$")
                st.bar_chart(pf, height=260, color="#4da3ff")
            with cg2:
                st.markdown("#### 🏆 Quem mais venceu (nº de itens)")
                qf = itens["vencedor"].value_counts().rename_axis("Fornecedor").rename("Itens")
                st.bar_chart(qf, height=260, color="#ff5d6c")

            # ----- variação de preço (subiu / caiu) -----
            st.markdown("#### 📈 Variação de preço (compara as 2 últimas datas)")
            datas = sorted(itens["data"].unique())
            if len(datas) < 2:
                st.info("Salve pelo menos **2 cotações em datas diferentes** pra ver o que subiu e o que caiu.")
            else:
                d_ant, d_atu = datas[-2], datas[-1]
                a = itens[itens["data"] == d_ant].groupby("codigo").agg(
                    produto=("produto", "first"), preco_ant=("preco_unit", "min")).reset_index()
                b = itens[itens["data"] == d_atu].groupby("codigo").agg(
                    produto=("produto", "first"), preco_atu=("preco_unit", "min")).reset_index()
                comp = pd.merge(a[["codigo", "preco_ant"]], b, on="codigo", how="inner")
                comp["Variação R$"] = (comp["preco_atu"] - comp["preco_ant"]).round(2)
                comp["Variação %"] = ((comp["preco_atu"] / comp["preco_ant"] - 1) * 100).round(1)
                comp = comp[comp["preco_ant"] != comp["preco_atu"]]

                st.caption(f"Comparando {d_ant.strftime('%d/%m/%Y')} → {d_atu.strftime('%d/%m/%Y')}")
                subiu = comp[comp["Variação R$"] > 0].sort_values("Variação %", ascending=False)
                caiu = comp[comp["Variação R$"] < 0].sort_values("Variação %")
                v1, v2, v3 = st.columns(3)
                cartao(v1, "c-rede", len(subiu), "Subiram de preço")
                cartao(v2, "c-emb", len(caiu), "Caíram de preço")
                cartao(v3, "c-fora", len(comp), "Mudaram no total")
                st.write("")

                def tabela_var(df):
                    t = df.rename(columns={"produto": "Produto", "preco_ant": "Antes", "preco_atu": "Agora"})
                    t = t[["Produto", "Antes", "Agora", "Variação R$", "Variação %"]].copy()
                    for c in ["Antes", "Agora", "Variação R$"]:
                        t[c] = t[c].apply(moeda)
                    return t

                cs, cc = st.columns(2)
                with cs:
                    st.markdown("**🔴 Maiores altas**")
                    st.dataframe(tabela_var(subiu.head(15)), use_container_width=True, hide_index=True)
                with cc:
                    st.markdown("**🟢 Maiores quedas**")
                    st.dataframe(tabela_var(caiu.head(15)), use_container_width=True, hide_index=True)

            # ----- gerenciar histórico -----
            with st.expander("🗂️ Cotações salvas (apagar se precisar)"):
                for _, c in cot_f.sort_values("data", ascending=False).iterrows():
                    cc1, cc2 = st.columns([4, 1])
                    cc1.write(f"**{c['data'].strftime('%d/%m/%Y')}** — salva em {c['criado_em']}")
                    if cc2.button("🗑️ Apagar", key=f"del_{c['id']}"):
                        apagar_cotacao(c["id"])
                        st.rerun()


# ---------------- CONFIG ----------------
with aba_cfg:
    st.subheader("Marcas cadastradas por categoria")
    st.write("Estas são as marcas fixas. Se aparecer marca 'não classificada', é porque ela não está aqui.")
    c1, c2, c3 = st.columns(3)
    cartao(c1, "c-fora", len(categorias["FORA"]), "Cotação FORA")
    cartao(c2, "c-emb", len(categorias["EMBREPAR"]), "Embrepar")
    cartao(c3, "c-rede", len(categorias["REDE"]), "Rede")
    st.write("")
    busca = st.text_input("Procurar uma marca")
    for nome in ["FORA", "EMBREPAR", "REDE"]:
        lst = categorias[nome]
        if busca:
            lst = [m for m in lst if normaliza(busca) in normaliza(m)]
        with st.expander(f"{nome} ({len(lst)})"):
            st.write(", ".join(sorted(lst)) if lst else "—")
